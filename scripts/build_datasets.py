from pathlib import Path
import shutil
import numpy as np
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
PROCESSED = ROOT / "data" / "processed"
EXPORTS = ROOT / "data" / "exports"
INFOGRAPHICS = ROOT / "assets" / "infographics"

DISCLAIMER = (
    "This project is an independent governance analytics and civic-tech research "
    "platform based on publicly available information."
)


def ensure_dirs() -> None:
    for path in [PROCESSED, EXPORTS, INFOGRAPHICS, ROOT / "reports", ROOT / "outputs"]:
        path.mkdir(parents=True, exist_ok=True)


def evidence_score(row: pd.Series) -> float:
    kind = str(row["evidence_type"]).lower()
    confidence = float(row["confidence"])
    if kind == "quantitative":
        base = 85
    elif kind in {"project_report", "regulator_context", "policy_target_context"}:
        base = 68
    elif kind == "qualitative":
        base = 55
    elif kind == "insufficient_public_evidence":
        return 0.0
    else:
        base = 45
    value_present = pd.notna(row.get("metric_value")) and str(row.get("metric_value")).strip() != ""
    metric_bonus = 10 if value_present else 0
    return round(min(100, base * confidence + metric_bonus), 2)


def build_category_scores(evidence: pd.DataFrame, weights: pd.DataFrame) -> pd.DataFrame:
    evidence = evidence.copy()
    evidence["evidence_score"] = evidence.apply(evidence_score, axis=1)
    evidence["is_scored"] = evidence["evidence_type"].ne("insufficient_public_evidence")

    category_grid = (
        evidence[["person"]]
        .drop_duplicates()
        .merge(weights[["category", "weight"]], how="cross")
    )
    grouped = (
        evidence[evidence["is_scored"]]
        .groupby(["person", "category"], as_index=False)
        .agg(
            category_score=("evidence_score", "mean"),
            evidence_count=("evidence_id", "count"),
            avg_confidence=("confidence", "mean"),
            quantitative_count=("evidence_type", lambda x: int((x == "quantitative").sum())),
        )
    )
    scores = category_grid.merge(grouped, on=["person", "category"], how="left")
    scores["category_score"] = scores["category_score"].fillna(0).round(2)
    scores["evidence_count"] = scores["evidence_count"].fillna(0).astype(int)
    scores["avg_confidence"] = scores["avg_confidence"].fillna(0).round(2)
    scores["quantitative_count"] = scores["quantitative_count"].fillna(0).astype(int)
    scores["weighted_points"] = (scores["category_score"] * scores["weight"]).round(2)
    scores["coverage_status"] = np.where(
        scores["evidence_count"] == 0,
        "insufficient_public_evidence",
        np.where(scores["quantitative_count"] > 0, "quantitative_or_mixed", "qualitative_only"),
    )
    return scores


def build_leader_scores(category_scores: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for person, group in category_scores.groupby("person"):
        total = group["weighted_points"].sum()
        evidence_count = group["evidence_count"].sum()
        quantitative_count = group["quantitative_count"].sum()
        coverage = (group["evidence_count"].gt(0).sum() / len(group)) * 100
        rows.append(
            {
                "person": person,
                "overall_governance_score": round(total, 2),
                "evidence_count": int(evidence_count),
                "quantitative_evidence_count": int(quantitative_count),
                "category_coverage_percent": round(coverage, 2),
                "interpretation": (
                    "Evidence-backed weighted score. Use with limitations; missing public data is not proof of poor performance."
                ),
                "disclaimer": DISCLAIMER,
            }
        )
    return pd.DataFrame(rows).sort_values("overall_governance_score", ascending=False)


def build_timeline(evidence: pd.DataFrame) -> pd.DataFrame:
    timeline = evidence.copy()
    timeline["metric_year"] = pd.to_numeric(timeline["metric_year"], errors="coerce")
    timeline = timeline.sort_values(["metric_year", "person", "category"])
    return timeline[
        [
            "metric_year",
            "person",
            "category",
            "indicator",
            "claim_summary",
            "source_id",
            "confidence",
        ]
    ]


def write_excel(
    evidence: pd.DataFrame,
    sources: pd.DataFrame,
    weights: pd.DataFrame,
    category_scores: pd.DataFrame,
    leader_scores: pd.DataFrame,
    timeline: pd.DataFrame,
) -> None:
    output = EXPORTS / "governance_analytics_dataset.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        leader_scores.to_excel(writer, sheet_name="Executive Scores", index=False)
        category_scores.to_excel(writer, sheet_name="Category Scores", index=False)
        evidence.to_excel(writer, sheet_name="Evidence Ledger", index=False)
        sources.to_excel(writer, sheet_name="Sources", index=False)
        weights.to_excel(writer, sheet_name="Scoring Weights", index=False)
        timeline.to_excel(writer, sheet_name="Timeline", index=False)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = load_workbook(output)
        header_fill = PatternFill("solid", fgColor="0B1F3A")
        gold_fill = PatternFill("solid", fgColor="C9A227")
        white_font = Font(color="FFFFFF", bold=True)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 55)
        ws = wb["Executive Scores"]
        for cell in ws[1]:
            cell.fill = gold_fill if cell.column <= 2 else header_fill
        wb.save(output)
    except Exception as exc:
        print(f"Excel styling skipped: {exc}")


def main() -> None:
    ensure_dirs()
    evidence = pd.read_csv(RAW / "evidence_ledger.csv")
    sources = pd.read_csv(RAW / "sources.csv")
    weights = pd.read_csv(RAW / "scoring_weights.csv")

    evidence = evidence.merge(sources[["source_id", "title", "organization", "url", "reliability_tier"]], on="source_id", how="left")
    evidence["evidence_score"] = evidence.apply(evidence_score, axis=1)
    evidence["source_available"] = evidence["url"].notna()
    evidence["data_quality_flag"] = np.select(
        [
            evidence["evidence_type"].eq("quantitative") & evidence["source_available"],
            evidence["evidence_type"].eq("insufficient_public_evidence"),
        ],
        ["source_backed_quantitative", "do_not_score_missing_data"],
        default="source_backed_contextual",
    )

    category_scores = build_category_scores(evidence, weights)
    leader_scores = build_leader_scores(category_scores)
    timeline = build_timeline(evidence)

    evidence.to_csv(PROCESSED / "clean_evidence_ledger.csv", index=False)
    sources.to_csv(PROCESSED / "source_verification_table.csv", index=False)
    category_scores.to_csv(PROCESSED / "category_scores.csv", index=False)
    leader_scores.to_csv(PROCESSED / "leader_scores.csv", index=False)
    timeline.to_csv(PROCESSED / "timeline.csv", index=False)
    weights.to_csv(PROCESSED / "scoring_weights.csv", index=False)

    # Tableau and Power BI can ingest these flat files directly.
    for file_name in [
        "clean_evidence_ledger.csv",
        "source_verification_table.csv",
        "category_scores.csv",
        "leader_scores.csv",
        "timeline.csv",
        "scoring_weights.csv",
    ]:
        shutil.copyfile(PROCESSED / file_name, EXPORTS / file_name)

    infographic = category_scores.pivot(index="person", columns="category", values="category_score").reset_index()
    infographic.to_csv(INFOGRAPHICS / "social_infographic_scorecard_data.csv", index=False)
    write_excel(evidence, sources, weights, category_scores, leader_scores, timeline)


if __name__ == "__main__":
    main()
